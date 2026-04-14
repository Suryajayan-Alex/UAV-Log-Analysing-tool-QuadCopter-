from __future__ import annotations

from datetime import datetime
from pathlib import Path
from textwrap import shorten

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle

from .branding import THEME
from .models import PlotResult, SignalStats


def _stats_frame(stats: list[SignalStats]) -> pd.DataFrame:
    rows = [
        {
            "Plot": stat.plot_title,
            "Signal": stat.signal,
            "Min": stat.minimum,
            "Max": stat.maximum,
            "Mean": stat.mean,
            "Samples": stat.samples,
        }
        for stat in stats
    ]
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["Plot", "Signal"], ignore_index=True)
    return df


def _hex_to_reportlab_rgb(color_hex: str) -> tuple[float, float, float]:
    value = color_hex.lstrip("#")
    return tuple(int(value[i : i + 2], 16) / 255.0 for i in (0, 2, 4))


def _hex_to_excel_rgb(color_hex: str) -> str:
    return color_hex.lstrip("#").upper()


def _style_excel_report(path: Path, metadata: dict[str, str], logo_path: Path | None = None) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = load_workbook(path)

        primary_fill = PatternFill(fill_type="solid", fgColor=_hex_to_excel_rgb(THEME.primary))
        secondary_fill = PatternFill(fill_type="solid", fgColor=_hex_to_excel_rgb(THEME.secondary))
        alt_fill = PatternFill(fill_type="solid", fgColor="F5F9FF")

        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="FFFFFF", bold=True, size=13)
        text_font = Font(color=_hex_to_excel_rgb(THEME.text_dark))
        bold_text_font = Font(color=_hex_to_excel_rgb(THEME.text_dark), bold=True)

        thin_side = Side(style="thin", color=_hex_to_excel_rgb(THEME.border))
        table_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        def style_cell(cell, *, fill=None, font=None, alignment=None, border=None, number_format=None):
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            if border is not None:
                cell.border = border
            if number_format is not None:
                cell.number_format = number_format

        if "Run_Info" in wb.sheetnames:
            info_ws = wb["Run_Info"]

            header_row = 1
            if logo_path and logo_path.exists():
                info_ws.insert_rows(1, amount=5)
                logo = XLImage(str(logo_path))
                logo.width = 220
                logo.height = 80
                info_ws.add_image(logo, "A1")

                info_ws["D2"] = "UAV Log Analysis"
                info_ws["D3"] = "Telemetry Report"
                info_ws["D2"].font = Font(size=20, bold=True, color=_hex_to_excel_rgb(THEME.primary))
                info_ws["D3"].font = Font(size=12, bold=True, color=_hex_to_excel_rgb(THEME.text_dark))
                header_row = 6

            info_ws.freeze_panes = f"A{header_row + 1}"
            info_ws.column_dimensions["A"].width = 24
            info_ws.column_dimensions["B"].width = 56

            for cell in info_ws[header_row]:
                if cell.value is None:
                    continue
                style_cell(
                    cell,
                    fill=primary_fill,
                    font=header_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

            for row in range(header_row + 1, info_ws.max_row + 1):
                style_cell(
                    info_ws[f"A{row}"],
                    font=bold_text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )
                style_cell(
                    info_ws[f"B{row}"],
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )

        if "Min_Max_Summary" in wb.sheetnames:
            summary_ws = wb["Min_Max_Summary"]

            summary_ws.merge_cells("A1:F1")
            summary_ws["A1"] = "Min/Max Summary by Plot Signal"
            style_cell(
                summary_ws["A1"],
                fill=primary_fill,
                font=title_font,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=table_border,
            )

            summary_ws.merge_cells("A2:F2")
            summary_ws["A2"] = (
                f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                f"Log: {metadata.get('log_file', '')}  |  Vehicle: {metadata.get('vehicle', '')}"
            )
            style_cell(
                summary_ws["A2"],
                fill=secondary_fill,
                font=header_font,
                alignment=Alignment(horizontal="left", vertical="center"),
                border=table_border,
            )

            summary_ws.freeze_panes = "A5"
            summary_ws.column_dimensions["A"].width = 40
            summary_ws.column_dimensions["B"].width = 34
            summary_ws.column_dimensions["C"].width = 14
            summary_ws.column_dimensions["D"].width = 14
            summary_ws.column_dimensions["E"].width = 14
            summary_ws.column_dimensions["F"].width = 12

            header_row = 4
            for cell in summary_ws[header_row]:
                if cell.value is None:
                    continue
                style_cell(
                    cell,
                    fill=primary_fill,
                    font=header_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

            data_start_row = 5
            data_end_row = summary_ws.max_row if summary_ws.max_row >= data_start_row else data_start_row - 1

            for row in range(data_start_row, data_end_row + 1):
                row_fill = alt_fill if (row - data_start_row) % 2 == 1 else None
                style_cell(
                    summary_ws[f"A{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )
                style_cell(
                    summary_ws[f"B{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )
                for col in ["C", "D", "E"]:
                    style_cell(
                        summary_ws[f"{col}{row}"],
                        fill=row_fill,
                        font=text_font,
                        alignment=Alignment(horizontal="right", vertical="center"),
                        border=table_border,
                        number_format="0.000",
                    )
                style_cell(
                    summary_ws[f"F{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="right", vertical="center"),
                    border=table_border,
                    number_format="0",
                )

            doc_start = max(data_end_row, header_row) + 2

            summary_ws.merge_cells(start_row=doc_start, start_column=1, end_row=doc_start, end_column=6)
            summary_ws.cell(row=doc_start, column=1, value="Header Documentation")
            style_cell(
                summary_ws.cell(row=doc_start, column=1),
                fill=secondary_fill,
                font=header_font,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=table_border,
            )

            doc_header_row = doc_start + 1
            summary_ws.cell(row=doc_header_row, column=1, value="Header")
            summary_ws.merge_cells(start_row=doc_header_row, start_column=2, end_row=doc_header_row, end_column=5)
            summary_ws.cell(row=doc_header_row, column=2, value="Description")
            summary_ws.cell(row=doc_header_row, column=6, value="Unit")

            for col in range(1, 7):
                style_cell(
                    summary_ws.cell(row=doc_header_row, column=col),
                    fill=primary_fill,
                    font=header_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

            docs = [
                ("Plot", "Chart name for the telemetry category.", "N/A"),
                ("Signal", "Specific plotted telemetry signal.", "Varies"),
                ("Min", "Minimum observed value for the signal.", "Signal unit"),
                ("Max", "Maximum observed value for the signal.", "Signal unit"),
                ("Mean", "Average value across all plotted samples.", "Signal unit"),
                ("Samples", "Count of valid samples used for stats.", "Count"),
            ]

            for idx, (name, desc, unit) in enumerate(docs, start=1):
                row = doc_header_row + idx
                summary_ws.cell(row=row, column=1, value=name)
                summary_ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                summary_ws.cell(row=row, column=2, value=desc)
                summary_ws.cell(row=row, column=6, value=unit)

                style_cell(
                    summary_ws.cell(row=row, column=1),
                    font=bold_text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )
                style_cell(
                    summary_ws.cell(row=row, column=2),
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                    border=table_border,
                )
                style_cell(
                    summary_ws.cell(row=row, column=6),
                    font=text_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

                for col in [3, 4, 5]:
                    style_cell(summary_ws.cell(row=row, column=col), border=table_border)

        if "Pass_Fail_Summary" in wb.sheetnames:
            pf_ws = wb["Pass_Fail_Summary"]

            pass_fill = PatternFill(fill_type="solid", fgColor="EAF8F1")
            fail_fill = PatternFill(fill_type="solid", fgColor="FDEEEF")
            na_fill = PatternFill(fill_type="solid", fgColor="EEF4FB")

            pf_ws.merge_cells("A1:E1")
            pf_ws["A1"] = "Pass/Fail Summary by Plot"
            style_cell(
                pf_ws["A1"],
                fill=primary_fill,
                font=title_font,
                alignment=Alignment(horizontal="center", vertical="center"),
                border=table_border,
            )

            pf_ws.merge_cells("A2:E2")
            pf_ws["A2"] = (
                f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
                f"Log: {metadata.get('log_file', '')}  |  Vehicle: {metadata.get('vehicle', '')}"
            )
            style_cell(
                pf_ws["A2"],
                fill=secondary_fill,
                font=header_font,
                alignment=Alignment(horizontal="left", vertical="center"),
                border=table_border,
            )

            pf_ws.freeze_panes = "A5"
            pf_ws.column_dimensions["A"].width = 40
            pf_ws.column_dimensions["B"].width = 13
            pf_ws.column_dimensions["C"].width = 14
            pf_ws.column_dimensions["D"].width = 14
            pf_ws.column_dimensions["E"].width = 74

            header_row = 4
            for cell in pf_ws[header_row]:
                if cell.value is None:
                    continue
                style_cell(
                    cell,
                    fill=primary_fill,
                    font=header_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

            data_start_row = 5
            data_end_row = pf_ws.max_row if pf_ws.max_row >= data_start_row else data_start_row - 1

            for row in range(data_start_row, data_end_row + 1):
                status = str(pf_ws[f"B{row}"].value or "").strip().upper()
                if status == "PASS":
                    row_fill = pass_fill
                elif status == "FAIL":
                    row_fill = fail_fill
                else:
                    row_fill = na_fill

                style_cell(
                    pf_ws[f"A{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center"),
                    border=table_border,
                )
                style_cell(
                    pf_ws[f"B{row}"],
                    fill=row_fill,
                    font=bold_text_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )
                style_cell(
                    pf_ws[f"C{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="right", vertical="center"),
                    border=table_border,
                    number_format="0",
                )
                style_cell(
                    pf_ws[f"D{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="right", vertical="center"),
                    border=table_border,
                    number_format="0",
                )
                style_cell(
                    pf_ws[f"E{row}"],
                    fill=row_fill,
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
                    border=table_border,
                )

        for sheet_name in ["Critical_Messages", "Skipped_Plots"]:
            if sheet_name not in wb.sheetnames:
                continue

            ws = wb[sheet_name]
            ws.freeze_panes = "A2"
            ws.column_dimensions["A"].width = 120

            for cell in ws[1]:
                if cell.value is None:
                    continue
                style_cell(
                    cell,
                    fill=primary_fill,
                    font=header_font,
                    alignment=Alignment(horizontal="center", vertical="center"),
                    border=table_border,
                )

            for row in range(2, ws.max_row + 1):
                style_cell(
                    ws[f"A{row}"],
                    font=text_font,
                    alignment=Alignment(horizontal="left", vertical="top", wrap_text=True),
                    border=table_border,
                )

        wb.save(path)
    except Exception:
        return


def export_excel_report(
    excel_path: Path,
    metadata: dict[str, str],
    plot_results: list[PlotResult],
    skipped_plots: list[str],
    critical_messages: list[str],
    logo_path: Path | None = None,
    cert_results: list[dict[str, str]] | None = None,
    current_stress: dict | None = None,
) -> None:
    stats = [stat for plot in plot_results for stat in plot.stats]
    stats_df = _stats_frame(stats)

    evaluations, _, overall_status = _evaluate_all_thresholds(plot_results, metadata=metadata)
    pass_plot_count = sum(1 for item in evaluations if item.status == "PASS")
    fail_plot_count = sum(1 for item in evaluations if item.status == "FAIL")
    na_plot_count = sum(1 for item in evaluations if item.status == "N/A")

    pass_fail_rows: list[dict[str, object]] = []
    for item in sorted(evaluations, key=lambda row: row.plot_title.lower()):
        failed_signals = [check.signal for check in item.checks if check.status == "FAIL"]
        if item.status == "N/A":
            notes = "No threshold-mapped signals in this plot"
        elif failed_signals:
            notes = "Failed signals: " + ", ".join(failed_signals)
        else:
            notes = "All mapped signals within limits"

        pass_fail_rows.append(
            {
                "Plot": item.plot_title,
                "Status": item.status,
                "Pass Signals": item.pass_count,
                "Fail Signals": item.fail_count,
                "Notes": notes,
            }
        )

    if not pass_fail_rows:
        pass_fail_rows.append(
            {
                "Plot": "No plots generated",
                "Status": "N/A",
                "Pass Signals": 0,
                "Fail Signals": 0,
                "Notes": "No threshold summary available.",
            }
        )

    pass_fail_df = pd.DataFrame(pass_fail_rows)

    info_rows = [
        {"Field": "Generated On", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Field": "Log File", "Value": metadata.get("log_file", "")},
        {"Field": "Vehicle", "Value": metadata.get("vehicle", "")},
        {"Field": "Pilot", "Value": metadata.get("pilot", "")},
        {"Field": "Co-Pilot", "Value": metadata.get("copilot", "")},
        {"Field": "Mission", "Value": metadata.get("mission", "")},
        {"Field": "Variant", "Value": metadata.get("variant", "Main")},
        {"Field": "Battery Capacity (mAh)", "Value": metadata.get("battery_capacity_mah", "")},
        {"Field": "Endurance Target (min)", "Value": metadata.get("endurance_minutes", "")},
        {"Field": "Observed Flight Time (min)", "Value": metadata.get("flight_time_minutes", "")},
        {"Field": "Endurance Status", "Value": metadata.get("endurance_status", "N/A")},
        {"Field": "Flight Start (UTC)", "Value": metadata.get("flight_utc", "N/A")},
        {"Field": "Weather Temp", "Value": metadata.get("weather_temp", "N/A")},
        {"Field": "Weather Wind", "Value": metadata.get("weather_wind", "N/A")},
        {"Field": "Weather Condition", "Value": metadata.get("weather_condition", "N/A")},
        {"Field": "Generated Plots", "Value": str(len(plot_results))},
        {"Field": "Overall Status", "Value": overall_status},
        {"Field": "Plots PASS", "Value": str(pass_plot_count)},
        {"Field": "Plots FAIL", "Value": str(fail_plot_count)},
        {"Field": "Plots N/A", "Value": str(na_plot_count)},
        {"Field": "Skipped Plots", "Value": str(len(skipped_plots))},
        {"Field": "Critical Messages", "Value": str(len(critical_messages))},
    ]
    info_df = pd.DataFrame(info_rows)

    critical_df = pd.DataFrame({"Critical Messages": critical_messages or ["No critical messages found"]})
    skipped_df = pd.DataFrame({"Skipped Plots": skipped_plots or ["None"]})

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        info_df.to_excel(writer, sheet_name="Run_Info", index=False)
        if cert_results:
            pd.DataFrame(cert_results).to_excel(writer, sheet_name="Variant_Certification", index=False)
        pass_fail_df.to_excel(writer, sheet_name="Pass_Fail_Summary", index=False, startrow=3)
        stats_df.to_excel(writer, sheet_name="Min_Max_Summary", index=False, startrow=3)
        critical_df.to_excel(writer, sheet_name="Critical_Messages", index=False)
        skipped_df.to_excel(writer, sheet_name="Skipped_Plots", index=False)

    _style_excel_report(excel_path, metadata=metadata, logo_path=logo_path)

def _fmt(value: float) -> str:
    return f"{value:.3f}"


def _draw_page_footer(c: canvas.Canvas, page_width: float) -> None:
    c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
    c.line(36, 30, page_width - 36, 30)
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 9)
    c.drawString(40, 18, "UAV Log Analyzer")
    c.drawRightString(page_width - 36, 18, f"Page {c.getPageNumber()}")


def _draw_page_header(c: canvas.Canvas, page_width: float, page_height: float, title: str, logo_path: Path | None = None) -> float:
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.primary))
    c.rect(0, page_height - 44, page_width, 44, stroke=0, fill=1)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(40, page_height - 28, title)

    if logo_path and logo_path.exists():
        try:
            logo = ImageReader(str(logo_path))
            lw, lh = logo.getSize()
            scale = min(120 / lw, 28 / lh)
            w = lw * scale
            h = lh * scale
            c.drawImage(logo, page_width - 42 - w, page_height - 36, width=w, height=h, mask="auto")
        except Exception:
            pass

    return page_height - 56


def _draw_cover_page(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    metadata: dict[str, str],
    plot_count: int,
    logo_path: Path | None = None,
) -> None:
    top_y = _draw_page_header(c, page_width, page_height, "UAV Flight Log Analysis Report", logo_path=logo_path)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 12)

    y = top_y - 18
    c.drawString(48, y, f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 22
    c.drawString(48, y, f"Log File: {metadata.get('log_file', '')}")
    y -= 22
    c.drawString(48, y, f"Air Vehicle: {metadata.get('vehicle', '')}")
    y -= 22
    c.drawString(48, y, f"Pilot: {metadata.get('pilot', '')}")
    y -= 22
    c.drawString(48, y, f"Co-Pilot: {metadata.get('copilot', '')}")
    y -= 22
    c.drawString(48, y, f"Mission: {metadata.get('mission', '')}")
    y -= 22
    c.drawString(48, y, f"Flight Start: {metadata.get('flight_utc', 'N/A')}")
    y -= 22
    c.drawString(48, y, f"Weather: {metadata.get('weather_temp', 'N/A')}, {metadata.get('weather_wind', 'N/A')}, {metadata.get('weather_condition', 'N/A')}")
    y -= 30

    c.setFont("Helvetica-Bold", 13)
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.primary))
    c.drawString(48, y, f"Generated Plots: {plot_count}")

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_stats_pages(c: canvas.Canvas, page_width: float, page_height: float, stats: list[SignalStats], logo_path: Path | None = None) -> None:
    if not stats:
        _draw_page_header(c, page_width, page_height, "Min/Max Summary by Plot Signal", logo_path=logo_path)
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 11)
        c.drawString(42, page_height - 90, "No plottable numeric signals were found.")
        _draw_page_footer(c, page_width)
        c.showPage()
        return

    ordered = sorted(stats, key=lambda s: (s.plot_title, s.signal))

    columns = [
        ("Plot", 42, 310, "left"),
        ("Signal", 310, 500, "left"),
        ("Min", 500, 578, "right"),
        ("Max", 578, 656, "right"),
        ("Mean", 656, 734, "right"),
        ("Samples", 734, page_width - 40, "right"),
    ]

    table_left = columns[0][1]
    table_right = columns[-1][2]
    row_height = 15

    def draw_header() -> float:
        _draw_page_header(c, page_width, page_height, "Min/Max Summary by Plot Signal", logo_path=logo_path)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 9)
        c.drawString(42, page_height - 70, "Columns are aligned and computed from valid plotted signal samples.")

        header_y = page_height - 94
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.secondary))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=0, fill=1)

        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _, _ in columns[1:]:
            c.line(x0, header_y, x0, header_y + row_height)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
        c.setFont("Helvetica-Bold", 9.5)
        for title, x0, x1, _ in columns:
            c.drawCentredString((x0 + x1) / 2, header_y + 4.2, title)

        return header_y - 1

    y_top = draw_header()

    for index, stat in enumerate(ordered):
        if y_top - row_height < 42:
            _draw_page_footer(c, page_width)
            c.showPage()
            y_top = draw_header()

        row_bottom = y_top - row_height
        if index % 2 == 1:
            c.setFillColorRGB(*_hex_to_reportlab_rgb("#F5F9FF"))
            c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=0, fill=1)

        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _, _ in columns[1:]:
            c.line(x0, row_bottom, x0, row_bottom + row_height)

        text_y = row_bottom + 4.0
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 8.7)

        values = [
            shorten(stat.plot_title, width=44, placeholder="..."),
            shorten(stat.signal, width=30, placeholder="..."),
            _fmt(stat.minimum),
            _fmt(stat.maximum),
            _fmt(stat.mean),
            str(stat.samples),
        ]

        for (title, x0, x1, align), value in zip(columns, values):
            del title
            if align == "left":
                c.drawString(x0 + 4, text_y, value)
            else:
                c.drawRightString(x1 - 4, text_y, value)

        y_top = row_bottom

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_text_list_page(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    title: str,
    lines: list[str],
    empty_text: str,
    logo_path: Path | None = None,
) -> None:
    _draw_page_header(c, page_width, page_height, title, logo_path=logo_path)
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 10)

    y = page_height - 86
    if not lines:
        c.drawString(42, y, empty_text)
    else:
        for line in lines:
            if y < 38:
                _draw_page_footer(c, page_width)
                c.showPage()
                _draw_page_header(c, page_width, page_height, title, logo_path=logo_path)
                c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
                c.setFont("Helvetica", 10)
                y = page_height - 86

            c.drawString(42, y, shorten(line, width=150, placeholder="..."))
            y -= 14

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_plot_pages(c: canvas.Canvas, page_width: float, page_height: float, plot_results: list[PlotResult], logo_path: Path | None = None) -> None:
    for plot in plot_results:
        _draw_page_header(c, page_width, page_height, plot.title, logo_path=logo_path)

        image = ImageReader(str(plot.image_path))
        image_width, image_height = image.getSize()

        usable_left = 20
        usable_right = page_width - 20
        usable_top = page_height - 52
        usable_bottom = 34

        max_width = usable_right - usable_left
        max_height = usable_top - usable_bottom
        scale = min(max_width / image_width, max_height / image_height)

        final_w = image_width * scale
        final_h = image_height * scale
        x = usable_left + (max_width - final_w) / 2
        y = usable_bottom + (max_height - final_h) / 2

        c.drawImage(image, x, y, width=final_w, height=final_h, preserveAspectRatio=True)

        _draw_page_footer(c, page_width)
        c.showPage()


def _draw_variant_certification_page(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    cert_results: list[dict[str, str]],
    metadata: dict[str, str],
    logo_path: Path | None = None,
) -> None:
    if not cert_results:
        return

    variant_type = metadata.get("variant", "Unknown")
    header_title = f"Variant Specific Criteria - {variant_type}"
    _draw_page_header(c, page_width, page_height, header_title, logo_path=logo_path)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica-Bold", 11)
    
    table_y = page_height - 90
    row_height = 18

    columns = [
        ("Criteria", 42, 280),
        ("Status", 280, 360),
        ("Measured", 360, 520),
        ("Expected", 520, page_width - 40)
    ]

    c.setFillColorRGB(*_hex_to_reportlab_rgb("#3B6EAF"))
    c.rect(columns[0][1], table_y, columns[-1][2] - columns[0][1], row_height, stroke=0, fill=1)
    
    c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
    c.rect(columns[0][1], table_y, columns[-1][2] - columns[0][1], row_height, stroke=1, fill=0)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
    c.setFont("Helvetica-Bold", 10)
    for title, x0, x1 in columns:
        c.drawCentredString((x0 + x1) / 2, table_y + 6, title)

    table_y -= row_height

    for i, res in enumerate(cert_results):
        if table_y < 35:
            _draw_page_footer(c, page_width)
            c.showPage()
            _draw_page_header(c, page_width, page_height, f"{header_title} (Cont.)", logo_path=logo_path)
            table_y = page_height - 90

        bg_color = "#FFFFFF" if res["Status"] == "PASS" else ("#FDEEEF" if res["Status"] == "FAIL" else "#FFFFFF")
        c.setFillColorRGB(*_hex_to_reportlab_rgb(bg_color))
        c.rect(columns[0][1], table_y, columns[-1][2] - columns[0][1], row_height, stroke=0, fill=1)
        
        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(columns[0][1], table_y, columns[-1][2] - columns[0][1], row_height, stroke=1, fill=0)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 9)
        
        c.drawString(columns[0][1] + 5, table_y + 6, shorten(res["Criterion"], width=40, placeholder="..."))
        
        status_color = THEME.success if res["Status"] == "PASS" else ("#B84545" if res["Status"] == "FAIL" else "#5A54A2")
        c.setFillColorRGB(*_hex_to_reportlab_rgb(status_color))
        c.setFont("Helvetica-Bold", 9)
        c.drawCentredString((columns[1][1] + columns[1][2]) / 2, table_y + 6, res["Status"])
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 9)
        c.drawCentredString((columns[2][1] + columns[2][2]) / 2, table_y + 6, shorten(res["Observed"], width=30, placeholder="..."))
        c.drawCentredString((columns[3][1] + columns[3][2]) / 2, table_y + 6, shorten(res["Condition"], width=40, placeholder="..."))

        table_y -= row_height

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_weather_dashboard_page(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    metadata: dict[str, str],
    plot_count: int,
    logo_path: Path | None = None,
) -> None:
    _draw_page_header(c, page_width, page_height, "Flight Conditions & Details", logo_path=logo_path)
    
    # Left Panel (Blue Sidebar)
    sidebar_w = 220
    sidebar_x = 40
    sidebar_y = 50
    sidebar_h = page_height - 130
    
    c.setFillColorRGB(*_hex_to_reportlab_rgb("#5A9AE6"))
    c.roundRect(sidebar_x, sidebar_y, sidebar_w, sidebar_h, 15, fill=1, stroke=0)
    
    # Left Panel Content
    c.setFillColorRGB(*_hex_to_reportlab_rgb("#FFFFFF"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(sidebar_x + 20, sidebar_y + sidebar_h - 40, "Flight Conditions")
    
    c.setFont("Helvetica", 10)
    c.drawString(sidebar_x + 20, sidebar_y + sidebar_h - 60, f"Armed Time: {metadata.get('armed_utc', 'N/A')}")
    
    temp_str = metadata.get("weather_temp", "N/A").replace(" °C", "°")
    c.setFont("Helvetica-Bold", 60)
    c.drawCentredString(sidebar_x + sidebar_w/2, sidebar_y + sidebar_h/2, temp_str)
    
    wind_str = metadata.get("weather_wind", "0")
    gust_str = metadata.get("weather_gust", "0")
    
    try:
        wind_val = float(wind_str.replace(" m/s", "").strip())
    except Exception:
        wind_val = 0.0
        
    try:
        gust_val = float(gust_str.replace(" m/s", "").strip())
    except Exception:
        gust_val = 0.0

    if wind_val <= 10.0 and gust_val <= 10.0:
        cond_str = "Good to fly"
        c.setFont("Helvetica", 16)
    else:
        cond_str = "Above the acceptable limit"
        c.setFont("Helvetica", 14)
        
    c.drawCentredString(sidebar_x + sidebar_w/2, sidebar_y + sidebar_h/2 - 40, cond_str)
    
    # Right Area - Cards
    right_x = sidebar_x + sidebar_w + 30
    right_w = page_width - right_x - 40
    
    c.setFillColorRGB(*_hex_to_reportlab_rgb("#2C3E50"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(right_x, sidebar_y + sidebar_h - 40, "Current Parameters")
    
    # Grid of Cards
    card_w = (right_w - 20) / 2
    card_h = 75
    
    cards = [
        ("Wind Speed", metadata.get("weather_wind", "N/A"), "Avg speed in m/s", "#3498db"),
        ("Gusts", metadata.get("weather_gust", "N/A"), "Max gust in m/s", "#e74c3c"),
        ("Cloud Cover", metadata.get("weather_cloud", "N/A"), "Sky obscurity %", "#95a5a6"),
        ("Precipitation", metadata.get("weather_precip", "N/A"), "Rainfall in mm", "#2980b9")
    ]
    
    start_y = sidebar_y + sidebar_h - 130
    for i, (title, val, desc, color) in enumerate(cards):
        row = i // 2
        col = i % 2
        cx = right_x + col * (card_w + 20)
        cy = start_y - row * (card_h + 20)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb("#FFFFFF"))
        c.setStrokeColorRGB(*_hex_to_reportlab_rgb("#E0E0E0"))
        c.roundRect(cx, cy, card_w, card_h, 8, fill=1, stroke=1)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb(color))
        c.circle(cx + 20, cy + card_h - 20, 5, fill=1, stroke=0)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb("#7F8C8D"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(cx + 35, cy + card_h - 24, title)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb("#2C3E50"))
        c.setFont("Helvetica-Bold", 20)
        c.drawCentredString(cx + card_w/2, cy + 25, val)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb("#BDC3C7"))
        c.setFont("Helvetica", 8)
        c.drawRightString(cx + card_w - 10, cy + 10, desc)

    # Flight Details below the cards
    details_y = 190
    c.setFillColorRGB(*_hex_to_reportlab_rgb("#2C3E50"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(right_x, details_y, "Flight Details")

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 11)

    y = details_y - 25
    col_1_x = right_x
    col_2_x = right_x + right_w / 2.0

    c.drawString(col_1_x, y, f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    c.drawString(col_1_x, y - 25, f"Log File: {metadata.get('log_file', '')}")
    c.drawString(col_1_x, y - 50, f"Air Vehicle: {metadata.get('vehicle', '')}")
    c.drawString(col_1_x, y - 75, f"Pilot: {metadata.get('pilot', '')}")

    c.drawString(col_2_x, y, f"Co-Pilot: {metadata.get('copilot', '')}")
    c.drawString(col_2_x, y - 25, f"Mission: {metadata.get('mission', '')}")
    c.drawString(col_2_x, y - 50, f"Flight Start: {metadata.get('flight_utc', 'N/A')}")
    c.drawString(col_2_x, y - 75, f"Weather: {metadata.get('weather_temp', 'N/A')}")

    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.primary))
    c.drawString(col_1_x, y - 105, f"Generated Plots: {plot_count}")

    _draw_page_footer(c, page_width)
    c.showPage()

def _draw_current_spike_page(
    c: canvas.Canvas, page_width: float, page_height: float, current_stress: dict, logo_path: Path | None = None
) -> None:
    results = current_stress.get("results", {})
    spike_events = results.get('spike_events', [])
    total_spikes = results.get('total_spikes', 0)
    
    # Matching Page 5 Min/Max Column styles
    columns = [
        ("ID", 42, 100),
        ("Start Time (s)", 100, 330),
        ("Duration (s)", 330, 560),
        ("Peak Current (A)", 560, page_width - 40),
    ]
    
    table_left = columns[0][1]
    table_right = columns[-1][2]
    row_height = 15
    
    def draw_header() -> float:
        _draw_page_header(c, page_width, page_height, "Current Over-Threshold Spike Analysis", logo_path=logo_path)
        
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 9)
        c.drawString(42, page_height - 70, f"Total Spikes Detected (>31A): {total_spikes}")
        
        header_y = page_height - 94
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.secondary))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=0, fill=1)
        
        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _ in columns[1:]:
            c.line(x0, header_y, x0, header_y + row_height)
            
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
        c.setFont("Helvetica-Bold", 9.5)
        for title, x0, x1 in columns:
            c.drawCentredString((x0 + x1) / 2, header_y + 4.2, title)
            
        return header_y - 1

    y_top = draw_header()
    
    if not spike_events:
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 8.7)
        c.drawString(42, y_top - 20, "No individual over-threshold spikes were detected.")
        _draw_page_footer(c, page_width)
        c.showPage()
        return

    for index, s in enumerate(spike_events):
        if y_top - row_height < 42:
            _draw_page_footer(c, page_width)
            c.showPage()
            y_top = draw_header()
            
        row_bottom = y_top - row_height
        if index % 2 == 1:
            c.setFillColorRGB(*_hex_to_reportlab_rgb("#F5F9FF"))
            c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=0, fill=1)
            
        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _ in columns[1:]:
            c.line(x0, row_bottom, x0, row_bottom + row_height)
            
        text_y = row_bottom + 4.0
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 8.7)
        
        values = [
            str(index + 1),
            f"{s['start']:.2f}",
            f"{s['duration']:.2f}",
            f"{s['max_current']:.1f}",
        ]
        
        # Center all values just like the professional summary
        for (title, x0, x1), value in zip(columns, values):
            c.drawCentredString((x0 + x1) / 2, text_y, value)
            
        y_top = row_bottom
        
    _draw_page_footer(c, page_width)
    c.showPage()


def export_pdf_report(
    pdf_path: Path,
    metadata: dict[str, str],
    plot_results: list[PlotResult],
    skipped_plots: list[str],
    critical_messages: list[str],
    logo_path: Path | None = None,
    cert_results: list[dict[str, str]] | None = None,
    current_stress: dict | None = None,
) -> None:
    stats = [stat for plot in plot_results for stat in plot.stats]
    pdf = canvas.Canvas(str(pdf_path), pagesize=landscape(A4))
    page_width, page_height = landscape(A4)

    # 1st Page: Variant Certification
    if cert_results:
        _draw_variant_certification_page(pdf, page_width, page_height, cert_results, metadata, logo_path=logo_path)
    
    # 2nd Page: Weather Dashboard
    _draw_weather_dashboard_page(pdf, page_width, page_height, metadata, len(plot_results), logo_path=logo_path)
    
    if current_stress:
        _draw_current_spike_page(pdf, page_width, page_height, current_stress, logo_path=logo_path)
    
    _draw_stats_pages(pdf, page_width, page_height, stats, logo_path=logo_path)


    if skipped_plots:
        _draw_text_list_page(
            pdf,
            page_width,
            page_height,
            "Skipped Plots",
            skipped_plots,
            empty_text="None",
            logo_path=logo_path,
        )

    _draw_text_list_page(
        pdf,
        page_width,
        page_height,
        "Critical Messages",
        critical_messages,
        empty_text="No critical messages found.",
        logo_path=logo_path,
    )
    _draw_plot_pages(pdf, page_width, page_height, plot_results, logo_path=logo_path)

    pdf.save()

# ---------- Threshold-based PASS/FAIL extension ----------
from dataclasses import dataclass
from typing import Any


@dataclass(frozen=True)
class _SignalThresholdCheck:
    signal: str
    minimum: float
    maximum: float
    mean: float
    metric: str
    lower_limit: float | None
    upper_limit: float | None
    source: str
    status: str


@dataclass(frozen=True)
class _PlotThresholdCheck:
    plot_key: str
    plot_title: str
    status: str
    checks: list[_SignalThresholdCheck]

    @property
    def pass_count(self) -> int:
        return sum(1 for item in self.checks if item.status == "PASS")

    @property
    def fail_count(self) -> int:
        return sum(1 for item in self.checks if item.status == "FAIL")


_STATUS_FAIL_COLOR = "#B84545"
_STATUS_NEUTRAL_COLOR = "#5A54A2"


def _status_fill(status: str) -> str:
    if status == "PASS":
        return THEME.success
    if status == "FAIL":
        return _STATUS_FAIL_COLOR
    return _STATUS_NEUTRAL_COLOR


def _status_row_bg(status: str) -> str:
    if status == "PASS":
        return "#EAF8F1"
    if status == "FAIL":
        return "#FDEEEF"
    return "#EEF4FB"


def _to_float_or_none(value: object) -> float | None:
    if value is None:
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if parsed == parsed else None


def _normalize_rule(rule: dict[str, Any]) -> dict[str, float | str] | None:
    metric = str(rule.get("metric", "range")).strip().lower()
    if metric not in {"range", "max", "mean", "min", "abs_max"}:
        metric = "range"

    lower = _to_float_or_none(rule.get("lower"))
    upper = _to_float_or_none(rule.get("upper"))

    if lower is not None and upper is not None and lower > upper:
        lower, upper = upper, lower

    if lower is None and upper is None:
        return None

    normalized: dict[str, float | str] = {"metric": metric}
    if lower is not None:
        normalized["lower"] = lower
    if upper is not None:
        normalized["upper"] = upper
    return normalized
def _normalize_limits_for_report(limit_key: str) -> tuple[float, float] | None:
    from .plot_dictionary import Y_AXIS_LIMITS

    raw = Y_AXIS_LIMITS.get(limit_key)
    if raw is None:
        return None

    try:
        lower = float(raw[0])
        upper = float(raw[1])
    except (TypeError, ValueError, IndexError):
        return None

    if lower == upper:
        pad = max(abs(lower) * 0.05, 1.0)
        lower -= pad
        upper += pad

    return (lower, upper) if lower < upper else (upper, lower)


def _acceptance_rule_for_signal(
    limit_key: str,
    signal_name: str,
    metadata: dict[str, str] | None = None,
) -> dict[str, float | str] | None:
    from .plot_dictionary import ACCEPTANCE_LIMITS

    def _variant_adjusted(rule: dict[str, Any] | None) -> dict[str, float | str] | None:
        normalized = _normalize_rule(rule) if isinstance(rule, dict) else None
        if normalized is None:
            return None

        if limit_key == "battery_cv_left":
            lower = _to_float_or_none((metadata or {}).get("battery_voltage_lower"))
            upper = _to_float_or_none((metadata or {}).get("battery_voltage_upper"))

            adjusted: dict[str, float | str] = {"metric": "range"}
            adjusted["lower"] = lower if lower is not None else _to_float_or_none(normalized.get("lower"))
            adjusted["upper"] = upper if upper is not None else _to_float_or_none(normalized.get("upper"))

            normalized_adjusted = _normalize_rule(adjusted)
            if normalized_adjusted is not None:
                return normalized_adjusted

        if limit_key == "battery_mah":
            capacity = _to_float_or_none((metadata or {}).get("battery_capacity_mah"))
            if capacity is not None and capacity > 0.0:
                return {
                    "metric": "max",
                    "lower": round(capacity * 0.80, 3),
                    "upper": round(capacity * 0.90, 3),
                }

        return normalized

    raw = ACCEPTANCE_LIMITS.get(limit_key)
    if raw is None or not isinstance(raw, dict):
        return None

    if "metric" in raw:
        return _variant_adjusted(raw)

    signal_lower = (signal_name or "").strip().lower()
    for key, value in raw.items():
        if key == "__default__":
            continue
        if not isinstance(value, dict):
            continue
        if str(key).strip().lower() == signal_lower:
            return _variant_adjusted(value)

    default_rule = raw.get("__default__")
    return _variant_adjusted(default_rule if isinstance(default_rule, dict) else None)


def _criteria_for_signal(
    limit_key: str,
    signal_name: str,
    metadata: dict[str, str] | None = None,
) -> tuple[dict[str, float | str] | None, str | None]:
    acceptance_rule = _acceptance_rule_for_signal(limit_key, signal_name, metadata=metadata)
    if acceptance_rule is not None:
        return acceptance_rule, "acceptance"

    fallback = _normalize_limits_for_report(limit_key)
    if fallback is None:
        return None, None

    return {"metric": "range", "lower": fallback[0], "upper": fallback[1]}, "y_axis"


def _evaluate_rule(stat: SignalStats, rule: dict[str, float | str]) -> tuple[str, str, float | None, float | None]:
    metric = str(rule.get("metric", "range")).strip().lower()
    lower = _to_float_or_none(rule.get("lower"))
    upper = _to_float_or_none(rule.get("upper"))

    if lower is not None and upper is not None and lower > upper:
        lower, upper = upper, lower

    if metric == "range":
        lower_ok = True if lower is None else stat.minimum >= lower
        upper_ok = True if upper is None else stat.maximum <= upper
        status = "PASS" if (lower_ok and upper_ok) else "FAIL"
        return status, metric, lower, upper

    if metric == "mean":
        value = stat.mean
    elif metric == "min":
        value = stat.minimum
    elif metric == "abs_max":
        value = max(abs(stat.minimum), abs(stat.maximum))
    else:
        value = stat.maximum
        metric = "max"

    lower_ok = True if lower is None else value >= lower
    upper_ok = True if upper is None else value <= upper
    status = "PASS" if (lower_ok and upper_ok) else "FAIL"
    return status, metric, lower, upper


def _criteria_label(metric: str, lower: float | None, upper: float | None) -> str:
    metric_name = {
        "range": "Min/Max",
        "max": "Max",
        "mean": "Mean",
        "min": "Min",
        "abs_max": "AbsMax",
    }.get(metric, "Metric")

    if lower is not None and upper is not None:
        return f"{metric_name} [{lower:.3f}, {upper:.3f}]"
    if upper is not None:
        return f"{metric_name} <= {upper:.3f}"
    if lower is not None:
        return f"{metric_name} >= {lower:.3f}"
    return metric_name


def _limit_key_for_signal(plot_key: str, signal_name: str) -> str:
    signal = (signal_name or "").lower()

    if plot_key == "battery_cv":
        if "voltage" in signal:
            return "battery_cv_left"
        if "current" in signal:
            return "battery_cv_right"

    if plot_key == "attitude":
        if "roll" in signal:
            return "attitude_roll"
        if "pitch" in signal:
            return "attitude_pitch"
        if "yaw" in signal:
            return "attitude_yaw"

    return plot_key


def _evaluate_plot_threshold(plot: PlotResult, metadata: dict[str, str] | None = None) -> _PlotThresholdCheck:
    checks: list[_SignalThresholdCheck] = []

    for stat in plot.stats:
        limit_key = _limit_key_for_signal(plot.key, stat.signal)
        rule, source = _criteria_for_signal(limit_key, stat.signal, metadata=metadata)
        if rule is None or source is None:
            continue

        status, metric, lower, upper = _evaluate_rule(stat, rule)
        checks.append(
            _SignalThresholdCheck(
                signal=stat.signal,
                minimum=stat.minimum,
                maximum=stat.maximum,
                mean=stat.mean,
                metric=metric,
                lower_limit=lower,
                upper_limit=upper,
                source=source,
                status=status,
            )
        )

    if not checks:
        plot_status = "N/A"
    elif any(item.status == "FAIL" for item in checks):
        plot_status = "FAIL"
    else:
        plot_status = "PASS"

    return _PlotThresholdCheck(plot_key=plot.key, plot_title=plot.title, status=plot_status, checks=checks)


def _evaluate_all_thresholds(
    plot_results: list[PlotResult],
    metadata: dict[str, str] | None = None,
) -> tuple[list[_PlotThresholdCheck], dict[str, _PlotThresholdCheck], str]:
    evaluations = [_evaluate_plot_threshold(plot, metadata=metadata) for plot in plot_results]

    by_key: dict[str, _PlotThresholdCheck] = {}
    for result in evaluations:
        by_key[result.plot_key] = result

    evaluable = [item for item in evaluations if item.status in {"PASS", "FAIL"}]
    if not evaluable:
        overall = "N/A"
    elif any(item.status == "FAIL" for item in evaluable):
        overall = "FAIL"
    else:
        overall = "PASS"

    return evaluations, by_key, overall


def _draw_cover_page_with_status(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    metadata: dict[str, str],
    plot_count: int,
    overall_status: str,
    evaluations: list[_PlotThresholdCheck],
    logo_path: Path | None = None,
) -> None:
    top_y = _draw_page_header(c, page_width, page_height, "UAV Flight Log Analysis Report", logo_path=logo_path)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 12)

    y = top_y - 18
    c.drawString(48, y, f"Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    y -= 22
    c.drawString(48, y, f"Log File: {metadata.get('log_file', '')}")
    y -= 22
    c.drawString(48, y, f"Air Vehicle: {metadata.get('vehicle', '')}")
    y -= 22
    c.drawString(48, y, f"Pilot: {metadata.get('pilot', '')}")
    y -= 22
    c.drawString(48, y, f"Co-Pilot: {metadata.get('copilot', '')}")
    y -= 22
    c.drawString(48, y, f"Mission: {metadata.get('mission', '')}")
    y -= 22
    c.drawString(48, y, f"Variant: {metadata.get('variant', 'Main')}")
    y -= 22
    c.drawString(48, y, f"Battery Capacity: {metadata.get('battery_capacity_mah', '')} mAh")
    y -= 22
    c.drawString(48, y, f"Endurance Target: {metadata.get('endurance_minutes', '')} min")
    y -= 22
    c.drawString(48, y, f"Observed Flight Time: {metadata.get('flight_time_minutes', '')} min ({metadata.get('endurance_status', 'N/A')})")
    y -= 24

    pass_count = sum(1 for item in evaluations if item.status == "PASS")
    fail_count = sum(1 for item in evaluations if item.status == "FAIL")
    na_count = sum(1 for item in evaluations if item.status == "N/A")

    c.setFont("Helvetica-Bold", 12)
    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.primary))
    c.drawString(48, y, f"Generated Plots: {plot_count}")

    status_box_y = y - 66
    status_box_h = 54
    c.setFillColorRGB(*_hex_to_reportlab_rgb(_status_fill(overall_status)))
    c.roundRect(48, status_box_y, 420, status_box_h, 10, stroke=0, fill=1)

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
    c.setFont("Helvetica-Bold", 17)
    c.drawString(64, status_box_y + 31, f"Overall Status: {overall_status}")

    c.setFont("Helvetica", 10)
    c.drawString(64, status_box_y + 14, f"Plots PASS: {pass_count}   FAIL: {fail_count}   N/A: {na_count}")

    c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
    c.setFont("Helvetica", 10)
    c.drawString(
        48,
        status_box_y - 16,
        "Threshold checks use acceptance limits first, then fallback to Y-axis limits.",
    )

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_pass_fail_summary_page(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    evaluations: list[_PlotThresholdCheck],
    overall_status: str,
    logo_path: Path | None = None,
) -> None:
    columns = [
        ("Plot", 42, 314, "left"),
        ("Status", 314, 384, "center"),
        ("Pass", 384, 444, "right"),
        ("Fail", 444, 504, "right"),
        ("Notes", 504, page_width - 40, "left"),
    ]

    row_height = 16
    table_left = columns[0][1]
    table_right = columns[-1][2]

    ordered = sorted(evaluations, key=lambda item: item.plot_title.lower())

    pass_count = sum(1 for item in ordered if item.status == "PASS")
    fail_count = sum(1 for item in ordered if item.status == "FAIL")
    na_count = sum(1 for item in ordered if item.status == "N/A")

    def draw_header() -> float:
        _draw_page_header(c, page_width, page_height, "Pass/Fail Summary", logo_path=logo_path)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(
            42,
            page_height - 72,
            f"Overall: {overall_status}   |   PASS: {pass_count}   FAIL: {fail_count}   N/A: {na_count}",
        )

        header_y = page_height - 96
        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.secondary))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=0, fill=1)

        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, header_y, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _, _ in columns[1:]:
            c.line(x0, header_y, x0, header_y + row_height)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.white))
        c.setFont("Helvetica-Bold", 9.5)
        for title, x0, x1, _ in columns:
            c.drawCentredString((x0 + x1) / 2, header_y + 4.5, title)

        return header_y - 1

    y_top = draw_header()

    for item in ordered:
        if y_top - row_height < 42:
            _draw_page_footer(c, page_width)
            c.showPage()
            y_top = draw_header()

        row_bottom = y_top - row_height

        c.setFillColorRGB(*_hex_to_reportlab_rgb(_status_row_bg(item.status)))
        c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=0, fill=1)

        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(THEME.border))
        c.rect(table_left, row_bottom, table_right - table_left, row_height, stroke=1, fill=0)
        for _, x0, _, _ in columns[1:]:
            c.line(x0, row_bottom, x0, row_bottom + row_height)

        failed_signals = [check.signal for check in item.checks if check.status == "FAIL"]
        if item.status == "N/A":
            note = "No threshold-mapped signals in this plot"
        elif failed_signals:
            note = "Failed: " + ", ".join(failed_signals)
        else:
            note = "All mapped signals within criteria"

        values = [
            shorten(item.plot_title, width=45, placeholder="..."),
            item.status,
            str(item.pass_count),
            str(item.fail_count),
            shorten(note, width=48, placeholder="..."),
        ]

        text_y = row_bottom + 4.3
        c.setFont("Helvetica", 8.7)

        for (title, x0, x1, align), value in zip(columns, values):
            del title
            if align == "left":
                c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
                c.drawString(x0 + 4, text_y, value)
            elif align == "right":
                c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
                c.drawRightString(x1 - 4, text_y, value)
            else:
                c.setFillColorRGB(*_hex_to_reportlab_rgb(_status_fill(value)))
                c.setFont("Helvetica-Bold", 8.8)
                c.drawCentredString((x0 + x1) / 2, text_y, value)
                c.setFont("Helvetica", 8.7)

        y_top = row_bottom

    _draw_page_footer(c, page_width)
    c.showPage()


def _draw_plot_pages_with_thresholds(
    c: canvas.Canvas,
    page_width: float,
    page_height: float,
    plot_results: list[PlotResult],
    evaluations_by_key: dict[str, _PlotThresholdCheck],
    logo_path: Path | None = None,
) -> None:
    for plot in plot_results:
        _draw_page_header(c, page_width, page_height, plot.title, logo_path=logo_path)

        image = ImageReader(str(plot.image_path))
        image_width, image_height = image.getSize()

        usable_left = 20
        usable_right = page_width - 20
        usable_top = page_height - 52
        usable_bottom = 80

        max_width = usable_right - usable_left
        max_height = usable_top - usable_bottom
        scale = min(max_width / image_width, max_height / image_height)

        final_w = image_width * scale
        final_h = image_height * scale
        x = usable_left + (max_width - final_w) / 2
        y = usable_bottom + (max_height - final_h) / 2

        c.drawImage(image, x, y, width=final_w, height=final_h, preserveAspectRatio=True)

        evaluation = evaluations_by_key.get(plot.key)
        status = evaluation.status if evaluation else "N/A"

        box_x = usable_left
        box_y = 40
        box_h = 30
        box_w = max_width

        c.setFillColorRGB(*_hex_to_reportlab_rgb(_status_row_bg(status)))
        c.setStrokeColorRGB(*_hex_to_reportlab_rgb(_status_fill(status)))
        c.roundRect(box_x, box_y, box_w, box_h, 8, stroke=1, fill=1)

        c.setFillColorRGB(*_hex_to_reportlab_rgb(_status_fill(status)))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(box_x + 10, box_y + 18, f"Threshold Check: {status}")

        c.setFillColorRGB(*_hex_to_reportlab_rgb(THEME.text_dark))
        c.setFont("Helvetica", 9)

        if evaluation is None or not evaluation.checks:
            detail = "No threshold mapping available for plotted signals."
        else:
            failed = [item.signal for item in evaluation.checks if item.status == "FAIL"]
            criteria = sorted({
                _criteria_label(item.metric, item.lower_limit, item.upper_limit) for item in evaluation.checks
            })
            if len(criteria) == 1:
                criteria_text = f" Criteria: {criteria[0]}"
            elif criteria:
                criteria_text = " Multiple criteria"
            else:
                criteria_text = ""

            if failed:
                detail = "Failed signals: " + ", ".join(failed) + criteria_text
            else:
                detail = f"Signals in range: {evaluation.pass_count}/{len(evaluation.checks)}" + criteria_text

        c.drawString(box_x + 180, box_y + 18, shorten(detail, width=90, placeholder="..."))

        _draw_page_footer(c, page_width)
        c.showPage()
